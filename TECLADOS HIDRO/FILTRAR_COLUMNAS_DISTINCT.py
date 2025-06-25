import openpyxl
from connect_db import connect

class DynamicExcelToDB:
    def __init__(self, config):
       
        self.config = config
        self.processed_combinations = set()
        self.inserted_records = []
        self.errors = []

    def process_excel_to_db(self):
        try:
            # Cargar Excel
            book = openpyxl.load_workbook(self.config['excel_path'])
            
            if self.config['sheet_name'] not in book.sheetnames:
                print(f"La hoja '{self.config['sheet_name']}' no existe")
                return []
            
            sheet = book[self.config['sheet_name']]
            cursor = connect()
            
            if cursor is None:
                print("No se pudo conectar a la base de datos")
                return []
            
            # Determinar rango de procesamiento
            max_row = self.config.get('max_rows') or sheet.max_row
            start_row = self.config.get('start_row', 2)
            
            print(f"Procesando desde fila {start_row} hasta {max_row}")
            print(f"Tabla padre: {self.config['parent_table']}")
            print(f"Tabla hija: {self.config['child_table']}")
            
            for row_num in range(start_row, max_row + 1):
                self._process_row(sheet, cursor, row_num)
            
            try:
                cursor.close()
            except:
                pass
            book.close()
            
            self._print_final_report()
            
            return self.inserted_records
            
        except Exception as e:
            print(f"Error general: {e}")
            return []

    def _process_row(self, sheet, cursor, row_num):
        try:
            # Obtener valores de Excel
            parent_value = sheet[f"{self.config['excel_parent_column']}{row_num}"].value
            child_value = sheet[f"{self.config['excel_child_column']}{row_num}"].value
            
            parent_value = self._clean_value(parent_value)
            child_value = self._clean_value(child_value)
            
            if not parent_value or not child_value:
                return
            
            # Verificar duplicados
            combination = (parent_value, child_value)
            if combination in self.processed_combinations:
                return
            
            self.processed_combinations.add(combination)
            
            # Buscar ID del registro padre
            parent_id = self._get_parent_id(cursor, parent_value)
            if not parent_id:
                error_msg = f"{self.config['parent_name_column']} '{parent_value}' no encontrado (fila {row_num})"
                self.errors.append(error_msg)
                print(f"  {error_msg}")
                return
            
            # Verificar si el registro hijo ya existe
            if self._child_exists(cursor, child_value, parent_id):
                print(f"  {self.config['child_name_column']} '{child_value}' ya existe")
                return
            
            # Insertar nuevo registro hijo
            if self._insert_child(cursor, child_value, parent_id):
                self.inserted_records.append({
                    'row': row_num,
                    'parent_name': parent_value,
                    'parent_id': parent_id,
                    'child_name': child_value
                })
                print(f" Fila {row_num}: '{child_value}' → {self.config['parent_name_column']} '{parent_value}' (ID: {parent_id})")
            
        except Exception as e:
            error_msg = f"Error en fila {row_num}: {str(e)}"
            self.errors.append(error_msg)
            print(f" {error_msg}")
            try:
                cursor.rollback()
            except:
                pass

    def _clean_value(self, value):
        if value is None or value == '-':
            return None
        return str(value).strip()

    def _get_parent_id(self, cursor, parent_value):
        query = f"SELECT {self.config['parent_id_column']} FROM {self.config['parent_table']} WHERE {self.config['parent_name_column']} = ?"
        cursor.execute(query, (parent_value,))
        result = cursor.fetchone()
        return result[0] if result else None

    def _child_exists(self, cursor, child_value, parent_id):
        query = f"""SELECT {self.config['child_id_column']} FROM {self.config['child_table']} 
                    WHERE {self.config['child_name_column']} = ? AND {self.config['child_parent_fk_column']} = ?"""
        cursor.execute(query, (child_value, parent_id))
        return cursor.fetchone() is not None

    def _insert_child(self, cursor, child_value, parent_id):
        try:
            query = f"""INSERT INTO {self.config['child_table']} ({self.config['child_name_column']}, {self.config['child_parent_fk_column']}) 
                        VALUES (?, ?)"""
            cursor.execute(query, (child_value, parent_id))
            cursor.commit()
            return True
        except Exception as e:
            print(f"Error al insertar: {e}")
            cursor.rollback()
            return False

    def _print_final_report(self):
        print(f"\n{'='*50}")
        print(f" PROCESO COMPLETADO")
        print(f" Registros insertados: {len(self.inserted_records)}")
        print(f" Errores encontrados: {len(self.errors)}")
        
        if self.errors:
            print(f"\n  ERRORES DETECTADOS:")
            for error in self.errors[:10]:
                print(f"   • {error}")
            if len(self.errors) > 10:
                print(f"   ... y {len(self.errors) - 10} errores más")


def process_custom_config(custom_config):
    processor = DynamicExcelToDB(custom_config)
    return processor.process_excel_to_db()

def create_config(excel_path, sheet_name, parent_table, parent_columns, child_table, child_columns, excel_columns, start_row=2):
    """
    
    parent_columns: {'id': 'IDTAXPHYLUM', 'name': 'TAXPHYLUM'}
    child_columns: {'id': 'IDTAXCLASE', 'name': 'TAXCLASE', 'fk': 'IDTAXPHILUM'}
    excel_columns: {'parent': 'B', 'child': 'C'}
    """
    return {
        'excel_path': excel_path,
        'sheet_name': sheet_name,
        'parent_table': parent_table,
        'parent_id_column': parent_columns['id'],
        'parent_name_column': parent_columns['name'],
        'child_table': child_table,
        'child_id_column': child_columns['id'],
        'child_name_column': child_columns['name'],
        'child_parent_fk_column': child_columns['fk'],
        'excel_parent_column': excel_columns['parent'],
        'excel_child_column': excel_columns['child'],
        'start_row': start_row,
        'max_rows': None
    }

if __name__ == "__main__":
    
    custom_config = create_config(
         excel_path=r'C:\Users\Chemilab\Downloads\Bases_datos_HB_marino.xlsx',
         sheet_name='721_PASTOS',
         parent_table='TAXGENERO',
         parent_columns={'id': 'IDTAXGENERO', 'name': 'TAXGENERO'},
         child_table='TAXESPECIE',
         child_columns={'id': 'IDTAXMORFOTIPO', 'name': 'TAXMORFOTIPO', 'fk': 'IDTAXGENERO'},
         excel_columns={'parent': 'F', 'child': 'A'}
     )
    result = process_custom_config(custom_config)